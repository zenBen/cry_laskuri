import pandas as pd
import sys
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from copy import copy
import forex_date as fd

def process_trades_for_laskuri(coin, file_path, out_path):
    """
    Processes trade data from a CSV file, filters it by a specific coin,
    and converts it to a format suitable for Laskuri tax reporting.

    Args:
        coin (str): The cryptocurrency to filter by (e.g., 'BTC', 'LTC', 'XMR').
        file_path (str): The path to the trades CSV file.

    Returns:
        str: The path to the processed CSV file.
    """
    try:
        print(f"Processing coin: {coin}, looking for file at: {file_path}")
        # Load the dataset
        dataset = pd.read_csv(file_path)

        # Ensure 'pair' column contains only strings and handle missing values
        dataset['pair'] = dataset['pair'].astype(str).fillna('')

        # Filter for rows where 'pair' includes the specified coin
        coin_data = dataset[dataset['pair'].str.contains(coin, na=False)]

        # Check if any rows contain the requested coin
        if coin_data.empty:
            print(f"No trades found for coin: {coin}. Skipping...")
            return None

        # Check if the currency pair includes fiat that is not EUR
        if not coin_data['pair'].str.contains('EUR').all():
            for idx, row in coin_data.iterrows():
                pair = row['pair']
                if 'EUR' not in pair:
                    # Extract the fiat currency (e.g., USD, GBP)
                    fiat_currency = pair.replace(coin, '').replace('/', '').strip()
                    target_datetime = pd.to_datetime(row['time'], errors='coerce').strftime('%Y-%m-%d %H:%M:%S')
                    if pd.isnull(target_datetime):
                        print(f"Invalid datetime format for row: {row}")
                        continue

                    # Get the forex rate for the fiat currency to EUR
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    with open(os.path.join(script_dir, '.fx_api_key'), 'r') as key_file:
                        api_key = key_file.read().strip()
                    forex_rate = fd.get_forex_rate_at_datetime(f"{fiat_currency}EUR", target_datetime, api_key)
                    if forex_rate is not None:
                        print(f"Exchange rate: {forex_rate:.4f}")

                    if forex_rate is not None:
                        # Convert the price and total to EUR
                        coin_data.at[idx, 'price'] = row['price'] * forex_rate
                        coin_data.at[idx, 'cost'] = row['cost'] * forex_rate
                    else:
                        print(f"Could not retrieve forex rate for {fiat_currency} to EUR at {target_datetime}.")
                        coin_data.at[idx, 'price'] = None
                        coin_data.at[idx, 'cost'] = None

        # Convert to the required format
        converted_data = pd.DataFrame({
            'AIKA - DATE/TIME': pd.to_datetime(coin_data['time']).dt.strftime('%d.%m.%Y %H:%M'),
            'TAPAHTUMA - EVENT': coin_data['type'].map({'buy': 'Osto', 'sell': 'Myynti'}),
            'MÄÄRÄ - AMOUNT': coin_data['vol'],
            'HINTA € / VIRTUAALIVALUUTTA - PRICE PER UNIT': coin_data['price'],
            'YHTEENSÄ - TOTAL': coin_data['cost'],
            'fee': coin_data['fee'],
            'LÄHDE - SOURCE': 'Kraken',
            'VIRTUAALIVALUUTTAA JÄLJELLÄ 1 - CURRENCY REMAINING 1': coin_data['vol'],
            'HANKINTAMENO TAI HANKINTAMENO-OLETTAMA/KULUTETTU VIRTUAALIVALUUTTA - PURCHASE COST OR DEEMED ACQ. COST': '',
            'DEEMED ACQ COST': '',
            'VOITTO /TAPPIO - PROFIT / LOSS': '',
            'VIRTUAALIVALUUTTAA JÄLJELLÄ 2 - CURRENCY REMAINING 2': '',
        })

        # Step 1: Convert decimals from period to comma
        converted_data['MÄÄRÄ - AMOUNT'] = converted_data['MÄÄRÄ - AMOUNT'].map(lambda x: f"{x:.8f}".replace('.', ','))
        converted_data['HINTA € / VIRTUAALIVALUUTTA - PRICE PER UNIT'] = converted_data['HINTA € / VIRTUAALIVALUUTTA - PRICE PER UNIT'].map(lambda x: f"{x:.2f} €".replace('.', ','))
        converted_data['YHTEENSÄ - TOTAL'] = converted_data['YHTEENSÄ - TOTAL'].map(lambda x: f"{x:.2f} €".replace('.', ','))

        # Convert AMOUNT back to numeric for calculations
        converted_data['amount_numeric'] = converted_data['MÄÄRÄ - AMOUNT'].str.replace(',', '.').astype(float)

        # Step 2: Calculate CURRENCY REMAINING correctly
        currency_remaining = 0
        currency_remaining_list = []

        for idx, row in converted_data.iterrows():
            event = row['TAPAHTUMA - EVENT']
            amount = row['amount_numeric']

            if event == 'Osto':
                currency_remaining += amount
            elif event == 'Myynti':
                currency_remaining -= amount

            currency_remaining_list.append(currency_remaining)

        converted_data['currency_remaining'] = currency_remaining_list

        # Step 3, 4 & 5: Add DEEMED ACQ COST explicitly, fix Purchase Cost and Profit/Loss
        purchase_queue = []

        for idx, row in converted_data.iterrows():
            event = row['TAPAHTUMA - EVENT']
            amount = row['amount_numeric']
            price_per_unit = float(row['HINTA € / VIRTUAALIVALUUTTA - PRICE PER UNIT'].replace(',', '.').replace(' €', ''))
            total = float(row['YHTEENSÄ - TOTAL'].replace(',', '.').replace(' €', ''))

            if event == 'Osto':
                purchase_queue.append({'remaining': amount, 'price': price_per_unit})
                converted_data.at[idx, 'HANKINTAMENO TAI HANKINTAMENO-OLETTAMA/KULUTETTU VIRTUAALIVALUUTTA - PURCHASE COST OR DEEMED ACQ. COST'] = ''
                converted_data.at[idx, 'DEEMED ACQ COST'] = ''
                converted_data.at[idx, 'VOITTO /TAPPIO - PROFIT / LOSS'] = ''
            elif event == 'Myynti':
                purchase_cost = 0
                remaining_amount = amount

                # obtain the sales fee
                fee = row['fee']

                # Apply FIFO purchase cost calculation
                while remaining_amount > 0 and purchase_queue:
                    purchase = purchase_queue[0]
                    used_amount = min(remaining_amount, purchase['remaining'])
                    purchase_cost += used_amount * purchase['price']
                    purchase['remaining'] -= used_amount
                    remaining_amount -= used_amount

                    if purchase['remaining'] <= 0:
                        purchase_queue.pop(0)

                # add fee to purchase cost becos fees aren't deductible from deemed acq cost
                cost_plus_fee = purchase_cost + fee

                deemed_acq_cost = price_per_unit * 0.2 * amount  # 20% deemed acquisition cost
                applicable_cost = max(cost_plus_fee, deemed_acq_cost)
                profit_or_loss = total - applicable_cost

                # Format values for concatenation
                purchase_cost_str = f"{purchase_cost:.2f} €".replace('.', ',')
                deemed_acq_cost_str = f"{deemed_acq_cost:.2f} €".replace('.', ',')

                # Parentheses logic and concatenation
                if cost_plus_fee < deemed_acq_cost:
                    combined_str = f"({purchase_cost_str}) / {deemed_acq_cost_str}"
                else:
                    combined_str = f"{purchase_cost_str} / ({deemed_acq_cost_str})"

                # Overwrite, then drop extra
                converted_data.at[idx, 'HANKINTAMENO TAI HANKINTAMENO-OLETTAMA/KULUTETTU VIRTUAALIVALUUTTA - PURCHASE COST OR DEEMED ACQ. COST'] = combined_str

                # Convert calculated values to strings with comma as decimal separator
                converted_data.at[idx, 'VOITTO /TAPPIO - PROFIT / LOSS'] = f"{profit_or_loss:.2f} €".replace('.', ',')
                converted_data.at[idx, 'DEEMED ACQ COST'] = f"{deemed_acq_cost:.2f} €".replace('.', ',')

        # Format CURRENCY REMAINING fields
        converted_data['VIRTUAALIVALUUTTAA JÄLJELLÄ 1 - CURRENCY REMAINING 1'] = converted_data['currency_remaining'].map(lambda x: f"{x:.8f}".replace('.', ','))
        converted_data['VIRTUAALIVALUUTTAA JÄLJELLÄ 2 - CURRENCY REMAINING 2'] = converted_data['currency_remaining'].map(lambda x: f"{x:.8f}".replace('.', ','))

        # Drop helper columns
        converted_data.drop(columns=['amount_numeric', 'currency_remaining', 'DEEMED ACQ COST'], inplace=True)

        # Save the final processed file
        processed_file_name = f'processed_trades_{coin}.csv'
        processed_file_path = os.path.join(out_path, processed_file_name)
        converted_data.to_csv(processed_file_path, index=False)

        return processed_file_path
    except FileNotFoundError:
        print(f"Error: File not found at path: {file_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def csv_to_xlsx_for_laskuri(coin, csv_file_name):
    """
    Converts a processed_trades_XXX.csv file to an XLSX file suitable for
    Verohallinto Laskuri, using the coin name in both filenames.

    Args:
        coin (str): The cryptocurrency (e.g., 'BTC', 'LTC', 'ETH') to process.
                    The corresponding CSV file should be named 'processed_trades_<coin>.csv'
                    and be in the current working directory.
    """
    try:
        # Load the CSV file
        df = pd.read_csv(csv_file_name)
        # Drop the 'fee' column
        if 'fee' in df.columns:
            df = df.drop(columns=['fee'])

        # Create the XLSX filename
        xlsx_file_name = f"vero_laskuri_{coin}.xlsx"

        # Load the template workbook
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_file = os.path.join(script_dir, "vero_laskuri_template.xlsx")
        if not os.path.exists(template_file):
            raise FileNotFoundError(f"Could not find the template file: {template_file}")
        wb = openpyxl.load_workbook(template_file)

        # Select the active sheet
        ws = wb.active

        # Insert the coin name in cell H10
        ws["H10"] = coin

        # Find the starting row for data insertion
        start_row = 16

        # Preserve formulas and styling in B4:E13
        preserved_data = {}
        for row in ws.iter_rows(min_row=4, max_row=13, min_col=2, max_col=5):
            for cell in row:
                preserved_data[cell.coordinate] = {'value': cell.value, 'formula': cell.value if cell.data_type == 'f' else None, 'font': copy(cell.font)}
                # preserved_data[cell.coordinate] = {'value': cell.value, 'formula': cell.formula if cell.has_style else None , 'font': copy(cell.font)}

        # Insert the DataFrame data into the sheet, starting at row 16
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Find the last row with data
        last_row = ws.max_row

        # update formulae in H, I and J columns
        for col in ['H', 'I', 'J']:
            for row in range(16, last_row + 1):  # Iterate through the rows where data is added
                cell = ws[f'{col}{row}'] # we get the cell as an openpyxl cell object
                if cell.data_type == 'f': # check if it is a formula
                    cell.value = f"={cell.value}"  # recalculate the formula
                    # cell.formula = f"={cell.formula}"  # recalculate the formula

        # update formula in L3
        formula_cell = "L3"
        ws[formula_cell] = f'=SUMIF(C16:C{last_row},E5,K16:K{last_row})'

        # Restoring the formulas and format
        for coord, data in preserved_data.items():
            cell = ws[coord]
            cell.value = data['value']
            # Restore the formula only if there was a formula to begin with
            if data['formula'] is not None:
              cell.value = data['formula']
            #   cell.formula = data['formula']
            cell.font = data['font']

        # Save the new workbook
        wb.save(os.path.join(os.path.dirname(csv_file_name), xlsx_file_name))

        print(f"Successfully converted '{csv_file_name}' to '{xlsx_file_name}'")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    if len(sys.argv) > 3:
        print("Usage: python trades_to_laskuri_xl.py <coin> [<file_path>]")
        print("  <coin>: The cryptocurrency to filter by (e.g., BTC, LTC, XMR).")
        print("  <file_path>: (Optional) The path to the trades CSV file. Defaults to current directory.")
        sys.exit(1)

    if len(sys.argv) == 3:
        file_path = sys.argv[2]
    else:
        file_path = os.path.join(os.getcwd(), "data", "trades.csv")  # Default to current directory/trades.csv

    if len(sys.argv) < 2:
        coins = ['BCH', 'BSV', 'BTC', 'ETC', 'ETH', 'LTC', 'REP', 'XLM', 'XMR', 'XRP', 'ZEC']
    else:
        coins = [sys.argv[1]]

    for coin in coins:  # Add a loop to process each coin
        csv_file_name = os.path.join(os.getcwd(), "output", f"processed_trades_{coin}.csv")
        if not os.path.exists(csv_file_name):
            if not os.path.exists(file_path):
                print(f"Error: File not found at path: {file_path}")
                sys.exit(1)
            result_path = process_trades_for_laskuri(coin, file_path, os.path.join(os.getcwd(), "output"))
            if result_path:
                print(f"Trades data converted to vero laskuri csv & xl and saved to: {result_path}")
        # raise FileNotFoundError(f"Could not find {csv_file_name} in the current directory.")

        csv_to_xlsx_for_laskuri(coin, csv_file_name)

